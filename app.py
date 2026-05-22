import datetime as dt
from datetime import timedelta 
import os
import openpyxl as op 
import pandas as pd
from openpyxl.utils.datetime import from_excel
import sys
from flask import Flask, render_template, request, jsonify
import webview
import tkinter as tk
from tkinter import filedialog

"""
NEXT STEP
-Auto calculate the total hours and prompt for a message in notes
-Allow the user to insert their transportation type (default to their usual)
-Change to use QR codes
CURRENT BUGS
"""










# CHANGE FOR EACH VERSION!!!!!!
currentBranch = "SK"

# Cell column numbers
CLOCKINNO = 3
CLOCKOUTNO = 4
TRANSPORTNO = 12
BRANCHNO = 14
NOTESNO = 18

# Help PyInstaller determine the base directory for resources
base_dir = '.'
if hasattr(sys, '_MEIPASS'):
    base_dir = os.path.join(sys._MEIPASS)

# APP DECLARATION
app = Flask(__name__,template_folder="templates")

# Collect screen dimension info
root = tk.Tk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.destroy()

# Create teacher's name list
teacherNames = []
teacherNameCount = 0

# Get current time and dates, store as variables
currentTime = dt.datetime.now()
#currentMonthYear = currentTime.strftime("%B") + " " + currentTime.strftime("%Y")
monthArray = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
monthArrayCaps = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
currentMonthYear = monthArrayCaps[currentTime.month - 1] + " " + str(currentTime.year)
# Set folder path for completed spreadsheets
yearRange = ""
if currentTime.month < 4:
    if currentTime.month == 3 and currentTime.day > 26:
        yearRange = str(currentTime.year) + "-" + str(currentTime.year + 1)
    else:
        yearRange = str(currentTime.year - 1) + "-" + str(currentTime.year)
else:
    yearRange = str(currentTime.year) + "-" + str(currentTime.year + 1)
namingConvention = " - Work Hours and Transportation - " + yearRange + ".xlsx"
folder_path = "G:/My Drive/Teachers Folder/Work Hours NEW FORMAT " + yearRange
# Get formatted date for searching through timesheets
formattedDate = currentTime.strftime("%d") + "-" + currentTime.strftime("%m") + "-" + currentTime.strftime("%Y")
stringFormat = "%d-%m-%Y"
stringDatetimeObj = dt.datetime.strptime(formattedDate, stringFormat)











# MAIN FUNCTIONS
def findTeacherNames():
    try:
        teacherNames.clear()
    except UnboundLocalError:
        print("TeacherNames array empty")
    global teacherNameCount 
    teacherNameCount = 0
    try:
        for item in os.listdir(folder_path):
            item_path = os.path.join(folder_path, item)
            if os.path.isfile(item_path):
                fileName = f"{item}"
                #Only read necessary files
                if fileName[len(fileName) - len(namingConvention): len(fileName)] == namingConvention:
                    teacherName = fileName[0:len(fileName) - (len(namingConvention))]
                    teacherNames.append(teacherName)
                    teacherNameCount = teacherNameCount + 1
    except FileNotFoundError:
        print("No directory")
def findUserWorkbook(userName):
    #After username is retrieved, copy teacher's excel file into openpyxl workbook, then store current sheet into a dataframe
    wb = op.load_workbook(f'{folder_path}/★{userName}{namingConvention}', data_only = False)
    return wb
def insertPunchTime(workSheet, clIn, clOut, brnch, notes, transport):
    # Define variables for cell search (time, rows, columns, etc.)
    maxRow = workSheet.max_row

    # Search for current date and fill in current time for punch clock
    for i in range (1, maxRow + 1):
        if workSheet.cell(row = i, column = 1).value == stringDatetimeObj:
            # Insert values
            workSheet.cell(row = i, column = CLOCKINNO).value = clIn # Clock in time
            workSheet.cell(row = i, column = CLOCKOUTNO).value = clOut # Clock out time
            workSheet.cell(row = i, column = BRANCHNO).value = brnch # Branch
            workSheet.cell(row = i, column = NOTESNO).value = notes # Notes
            if transport != "":
                workSheet.cell(row = i, column = TRANSPORTNO).value = transport # Transportation
            return True
        else:
            # Check if value is a datetime that got converted into a serial number
            excel_serial_date = workSheet.cell(row = i, column = 1).value
            python_datetime = ""
            if isinstance(excel_serial_date, (int, float)):
                python_datetime = from_excel(excel_serial_date)
            if python_datetime == stringDatetimeObj:
                # Insert values
                workSheet.cell(row = i, column = CLOCKINNO).value = clIn # Clock in time
                workSheet.cell(row = i, column = CLOCKOUTNO).value = clOut # Clock out time
                workSheet.cell(row = i, column = BRANCHNO).value = brnch # Branch
                workSheet.cell(row = i, column = NOTESNO).value = notes # Notes
                if transport != "":
                    workSheet.cell(row = i, column = TRANSPORTNO).value = transport # Transportation
                return True
    return False
def findExcelFileInFolder(folder_path, namingConvention):
    # Copy teacher's excel file into openpyxl workbook, then store current sheet into a dataframe
    wb = op.load_workbook(f'{folder_path}/{namingConvention}.xlsx', data_only = True)
    return wb
def checkTotalWorkTime(clockInTime):
    clockInHour = ""
    # Turn clock in time into useable integer, or just the hour if time object
    if type(clockInTime).__name__ == "str":
        for char in clockInTime:
            if char == ":":
                break
            else:
                clockInHour = clockInHour + char
        clockInHour = int(clockInHour)
    else:
        clockInHour = clockInTime.hour

    # Calculate total hours worked minus 1 for days with a lunch hour
    totalHoursWorked = currentTime.hour - clockInHour
    if clockInHour < 12 and currentTime.hour > 12:
        totalHoursWorked = totalHoursWorked - 1
    return totalHoursWorked
def checkRecentData(workSheet):
    # Define variables for cell search (time, rows, columns, etc.)
    maxRow = workSheet.max_row
    formattedDate = currentTime.strftime("%d") + "-" + currentTime.strftime("%m") + "-" + currentTime.strftime("%Y")
    stringFormat = "%d-%m-%Y"
    stringDatetimeObj = dt.datetime.strptime(formattedDate, stringFormat)

    # Search for current date and fill in current time for punch clock
    for i in range (1, maxRow + 1):
        if workSheet.cell(row = i, column = 1).value == stringDatetimeObj:
            recentDataObject = {
                # Retrieve values
                "clockin": str(workSheet.cell(row = i, column = CLOCKINNO).value),
                "clockout": str(workSheet.cell(row = i, column = CLOCKOUTNO).value),
                "branch": str(workSheet.cell(row = i, column = BRANCHNO).value),
                "notes": str(workSheet.cell(row = i, column = NOTESNO).value),
                "transport": str(workSheet.cell(row = i, column = TRANSPORTNO).value)
            }
            for item in recentDataObject:
                if recentDataObject[item] == "None":
                    recentDataObject[item] = ""
            return recentDataObject
        else:
            # Check if value is a datetime that got converted into a serial number
            excel_serial_date = workSheet.cell(row = i, column = 1).value
            python_datetime = ""
            if isinstance(excel_serial_date, (int, float)):
                python_datetime = from_excel(excel_serial_date)
            if python_datetime == stringDatetimeObj:
                recentDataObject = {
                # Retrieve values
                "clockin": str(workSheet.cell(row = i, column = CLOCKINNO).value),
                "clockout": str(workSheet.cell(row = i, column = CLOCKOUTNO).value),
                "branch": str(workSheet.cell(row = i, column = BRANCHNO).value),
                "notes": str(workSheet.cell(row = i, column = NOTESNO).value),
                "transport": str(workSheet.cell(row = i, column = TRANSPORTNO).value)
                }
                for item in recentDataObject:
                    if recentDataObject[item] == "None":
                        recentDataObject[item] = ""
                return recentDataObject
    failedDataObject = {
                "clockin": "",
                "clockout": "",
                "branch": "",
                "notes": "",
                "transport": ""
            }
    return failedDataObject
def getWorksheet():
    # Change the sheet depending on whether we're past payday
    sheetToUse = ""
    if currentTime.day > 25:
        if currentTime.month == 12:
             sheetToUse = monthArrayCaps[0] + " " + currentTime.strftime("%Y")
        else:
            sheetToUse = monthArrayCaps[currentTime.month] + " " + currentTime.strftime("%Y")
    else:
        sheetToUse = monthArrayCaps[currentTime.month - 1] + " " + currentTime.strftime("%Y")
    return str(sheetToUse)

# EXTRA FUNCTIONS
def create():
    myWindowWidth = screen_width - 100
    myWindowHeight = screen_height - 100
    xLocation = (screen_width / 2) - (myWindowWidth / 2)
    yLocation = (screen_height / 2) - (myWindowHeight / 2)
    return webview.create_window('Handz Punch Clock', app, width=myWindowWidth, height=myWindowHeight, x=xLocation, y=yLocation)
def openFolderDialog():
    # Hide the main Tkinter window
    root = tk.Tk()
    root.withdraw()

    # Open the folder selection dialog
    userFolderPath = filedialog.askdirectory(
        title="Select a folder",
        initialdir=folder_path  # You can set an initial directory
    )

    if userFolderPath:
        root.destroy()
        return userFolderPath
    else:
        root.destroy()
        userFolderPath = folder_path
        return userFolderPath










# Startup
@app.route("/")
def startup():
    findTeacherNames()
    return render_template('index.html', teacherNames=teacherNames, folder_path=folder_path, 
                           teacherNameCount=teacherNameCount, currentBranch=currentBranch)
    
# Common routes
@app.route('/update', methods=['POST'])
def update():
    # Get data from javascript page
    data = request.get_json()
    scannedName = data.get('owner')
    clockInTime = data.get('var1')
    clockOutTime = data.get('var2')
    branch = data.get('var3')
    notes = data.get('var4')
    transport = data.get('var5')
    print(f"Transport is: {transport}")
    # Find user workbook and error check
    userWorkbook = findUserWorkbook(scannedName)
    if userWorkbook is False:
        return "File not found"
    # Find user worksheet and error check
    userWorksheet = getWorksheet()
    if userWorksheet is False:
        return "Worksheet within file not found"
    # Attempt to insert punch time, return if failed
    if not insertPunchTime(userWorkbook[userWorksheet], clockInTime, clockOutTime, branch, notes, transport):
        return "Failed to insert data"
    # Attempt to save workbook, return if failed
    try:
        userWorkbook.save(f"{folder_path}/★{scannedName}{namingConvention}")
        return "success"
    except PermissionError:
        return "PermissionError"
    except Exception as e:
        return f"{e}" 
@app.route("/changeDirectory", methods=['POST'])
def changeDirectory():
    global folder_path 
    folder_path = openFolderDialog()
    findTeacherNames()
    mainPageData = {
        "fp": folder_path,
        "tn": teacherNames,
        "tnc": teacherNameCount
    }
    return jsonify(mainPageData)
@app.route("/collectFileData", methods=['POST'])
def collectFileData():
    data = request.get_json()
    userWorkbook = findUserWorkbook(data)
    userWorksheet = getWorksheet()
    recentData = checkRecentData(userWorkbook[userWorksheet])
    return jsonify(recentData)

if __name__ == '__main__':
    #app.run(debug=True)
    window = create()
    webview.start()